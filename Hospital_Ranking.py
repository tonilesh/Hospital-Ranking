import urllib
import urllib.request
import zipfile
import os
import sqlite3
import string
import csv
import openpyxl
import math

sqlite_file = "medicare_hospital_compare.db"
directory = "staging"
zip_file_name = 'Hospital_Revised_Flatfiles.zip'
hospital_excel_file = "hospital_ranking_focus_states.xlsx"
temp_national_rank = "temp_national_rank"
temp_focus_state = "temp_focus_state"

class StdevFunc:
    def __init__(self):
        self.M = 0.0
        self.S = 0.0
        self.k = 1

    def step(self, value):
        try:
            # automatically convert text to float, like the rest of SQLite
            val = float(value)  # if fails, skips this iteration, which also ignores nulls
            tM = self.M
            self.k += 1
            self.M += ((val - tM) / self.k)
            self.S += ((val - tM) * (val - self.M))
        except:
            pass

    def finalize(self):
        if self.k < 3:
            return None
        return math.sqrt(self.S / (self.k-2))

# function to create staging directory
def createstagingdirectoty():
    path = os.path.join(directory)
    try:
        if not os.path.isdir(directory):
            os.mkdir(path)
    except IOError as error:
        print(error)
    else:
        print("Staging directory created successfully.")


# unzip downloaded zip file and copy all csv into it to staging
def unzip_hospital_file():
    try:
        outputFilename = "Hospital_Revised_Flatfiles.zip"
        zfobj = zipfile.ZipFile(outputFilename)
        for name in zfobj.namelist():
            uncompressed = zfobj.read(name)
            # save uncompressed data to disk
            outputFilename = "staging/" + name
            print("Saving extracted file to ", outputFilename)
            output = open(outputFilename, 'wb')
            output.write(uncompressed)
            output.close()
    except IOError as error:
        print(error)


# download Hospital zip file in local working directory
def download_Hospital_Zip_File():
    try:
        url = "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1"
        url += '?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip'

        response = urllib.request.urlopen(url)
        # response = urllib.urlopen(url)
        zippedData = response.read()
        output = open(zip_file_name, 'wb')
        output.write(zippedData)
        output.close()
    except RuntimeError as error:
        print(error)


# download excel file in working directory
def download_Hospital_Ranking_xlsx():
    url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
    urllib.request.urlretrieve(url, hospital_excel_file)

#function to read hospital_ranking excel file and generate sql statments
def read_Excel_RankingSheet():
    wb = openpyxl.load_workbook(hospital_excel_file)
    sheet_1_name = wb.get_sheet_names()[0]
    sheet_1 = wb.get_sheet_by_name(sheet_1_name)
    cleanedrows = []
    i = 2
    while sheet_1.cell(row=i, column=1).value != None:
        excelrows = []
        val = sheet_1.cell(row=i, column=1).value
        excelrows.append(val)
        val = str(sheet_1.cell(row=i, column=2).value)
        excelrows.append(val)
        i += 1
        cleanedrows.append(excelrows)

    #print(cleanedrows)
    return cleanedrows

#function to read states from focus list
def read_Excel_Focus_States():
    wb = openpyxl.load_workbook(hospital_excel_file)
    sheet_2_name = wb.get_sheet_names()[1]
    sheet_2 = wb.get_sheet_by_name(sheet_2_name)
    cleanedrows = []
    stateDict = {};
    i = 2;
    while sheet_2.cell(row=i, column=1).value != None:
        #excelrows = []
        #val = sheet_2.cell(row=i, column=1).value
        #excelrows.append(val)
        val = str(sheet_2.cell(row=i, column=2).value)
        cleanedrows.append(val)
        stateDict[val] = str(sheet_2.cell(row=i, column=1).value)
        i += 1

    return stateDict

def create_TempTable_For_Rank_And_State(cur):
    temprank = "create table if not exists " +  temp_national_rank + "(providerid text,ranking integer)"
    cur.execute(temprank)

def insert_Records_Into_Temp_Rank_Focus():
    temprankrows = read_Excel_RankingSheet()
    tempstaterows = read_Excel_Focus_States()
    #print(temprankrows)
    #print(tempstaterows)
    conn = sqlite3.connect(sqlite_file)
    conn.create_aggregate("stdev", 1, StdevFunc)
    cur = conn.cursor()
    create_TempTable_For_Rank_And_State(cur)
    try:
        stmt = "INSERT INTO " + temp_national_rank + " VALUES(?,?); "
        #print(stmt)
        cur.executemany(stmt,temprankrows)
        wb = openpyxl.Workbook()
        import_National_Ranking_Into_DB(cur,wb)
        export_Rows_To_Excel_State_Ranking(tempstaterows, cur,wb)
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wb.save("hospital_ranking.xlsx")
        wb = openpyxl.Workbook()
        create_Measure_Statistics_Excel_Nationwide_State(tempstaterows, cur, wb)
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        wb.save("measure_statistics.xlsx")
        conn.commit()
        conn.close()

    except RuntimeError as error:
        print(error)

def import_National_Ranking_Into_DB(cur,wb):
    sql_str =  "Select provider_id,hospital_name,city,state,county_name,tnr.ranking from hospital_general_information hgi"
    sql_str += " inner join temp_national_rank tnr on hgi.provider_id = tnr.providerid"
    sql_str += " order by tnr.ranking limit 100"

    rows = cur.execute(sql_str)
    export_Rows_To_Excel_Nationl_Ranking(rows,wb)

def export_Rows_To_Excel_Nationl_Ranking(rows,wb):
    if (os.path.isfile("hospital_ranking.xlsx")):
        os.remove("hospital_ranking.xlsx")
    create_hospital_ranking_excel("Nationwide",rows,wb)


def export_Rows_To_Excel_State_Ranking(staterows,cur,wb):
    sql_str = "";
    for staterow in staterows:
        print(staterow)
        sql_str =  "Select provider_id,hospital_name,city,state,county_name,tnr.ranking from hospital_general_information hgi"
        sql_str += " inner join temp_national_rank tnr on hgi.provider_id = tnr.providerid and hgi.state='" + staterow + "'"
        sql_str += " order by tnr.ranking limit 100"
        rows = cur.execute(sql_str)
        create_hospital_ranking_excel(staterow, rows,wb)

def create_hospital_ranking_excel(sheetname,rows,wb):
    sheet_1 = wb.create_sheet(sheetname)

    sheet_1.cell(row=1, column=1, value="Provider Id")
    sheet_1.cell(row=1, column=2, value="Hospital Name")
    sheet_1.cell(row=1, column=3, value="City")
    sheet_1.cell(row=1, column=4, value="State")
    sheet_1.cell(row=1, column=5, value="County")

    i = 2;
    for row in rows:
        sheet_1.cell(row=i, column=1, value=row[0])
        sheet_1.cell(row=i, column=2, value=row[1])
        sheet_1.cell(row=i, column=3, value=row[2])
        sheet_1.cell(row=i, column=4, value=row[3])
        sheet_1.cell(row=i, column=5, value=row[4])
        i += 1


def create_Measure_Statistics_Excel_Nationwide_State(states,cur,wb):
    sql_str = "";
    if (os.path.isfile("measure_statistics.xlsx")):
        os.remove("measure_statistics.xlsx")
    sql_str = "select measure_id,measure_name,min(cast(score as integer)) as minimum,max(cast(score as integer)) as maximum,avg(cast(score as inetger)) as average,"
    sql_str += " stdev(cast(score as intger)) as standarddeviation from timely_and_effective_care___hospital  where "
    sql_str += "score not glob '*[A-Za-z]*' group by measure_name"
    rows = cur.execute(sql_str)
    create_Measure_Statistics_Excel(rows,"Nationwide",wb)


    for state in states:
        sql_str = "select measure_id,measure_name,min(cast(score as integer)) as minimum,max(cast(score as integer)) as maximum,avg(cast(score as inetger)) as average,"
        sql_str += "stdev(cast(score as intger)) as standarddeviation from timely_and_effective_care___hospital "
        sql_str += "where score not glob '*[A-Za-z]*' and state='" + state +"' group by measure_name"
        rows = cur.execute(sql_str)
        print(states.get(state,"none"))
        create_Measure_Statistics_Excel(rows,states.get(state,"none"),wb)


def create_Measure_Statistics_Excel(rows,sheetname,wb):
    sheet_1 = wb.create_sheet(sheetname)

    sheet_1.cell(row=1, column=1, value="Measure Id")
    sheet_1.cell(row=1, column=2, value="Measure Name")
    sheet_1.cell(row=1, column=3, value="Minimum")
    sheet_1.cell(row=1, column=4, value="Maximum")
    sheet_1.cell(row=1, column=5, value="Average")
    sheet_1.cell(row=1, column=6, value="Standard Deviation")

    i = 2;
    for row in rows:
        sheet_1.cell(row=i, column=1, value=row[0])
        sheet_1.cell(row=i, column=2, value=row[1])
        sheet_1.cell(row=i, column=3, value=row[2])
        sheet_1.cell(row=i, column=4, value=row[3])
        sheet_1.cell(row=i, column=5, value=row[4])
        sheet_1.cell(row=i, column=5, value=row[5])
        i += 1

# create sqlite db
def create_Sqlite_Db():
    if (os.path.isfile(sqlite_file)):
        os.remove(sqlite_file)
    return sqlite3.connect(sqlite_file)


# import data from csv to DB
def csvToDb():
    path = os.path.join(directory)
    conn = create_Sqlite_Db()
    cur = conn.cursor()
    alpha = string.ascii_letters
    for root, dirs, files in os.walk(directory):
        for csvFile in files:
            if (csvFile.endswith(".csv") and csvFile != "FY2015_Percent_Change_in_Medicare_Payments.csv"):
                print(csvFile)
                print(directory + "/" + csvFile)
                with open(directory + "/" + csvFile, mode='r', encoding="ISO-8859-1") as fin:
                    fin.seek(0)
                    reader = csv.DictReader(fin)

                    # Keep the order of the columns name just as in the CSV
                    fields = reader.fieldnames
                    cols = []

                    # Set field and type
                    for f in fields:
                        fieldname = clean_Table_Column_Names(f, "c", alpha)
                        cols.append("%s" % (fieldname))

                    csvFile = clean_Table_Column_Names(csvFile, "t", alpha)
                    # Generate create table statement:
                    stmt = "CREATE TABLE " + csvFile + " (%s)" % ",".join(cols)
                    #print(stmt)
                    # con = sqlite3.connect(":memory:")
                    # cur = con.cursor()
                    cur.execute(stmt)
                    print("Table created")
                    fin.seek(0)

                    reader = csv.reader(escapingGenerator(fin))

                    # skip header
                    next(reader, None)
                    cleanedrows = [];
                    for row in reader:
                        # below condition will take care of handling empty rows which doesn't contain exact number of columns
                        # as number of columns in header.
                        if (len(row) == len(cols)):
                            cleanedrows.append(row)

                    #print(cleanedrows)
                    # Generate insert statement:
                    stmt = "INSERT INTO " + csvFile + " VALUES(%s);" % ','.join('?' * len(cols))
                    print(stmt)
                    print(cleanedrows)
                    cur.executemany(stmt, cleanedrows)
                    conn.commit()

    cur.close()
    del cur
    conn.close()
    return conn


def escapingGenerator(f):
    for line in f:
        yield line.encode("ascii", "xmlcharrefreplace").decode("ascii")


def clean_Table_Column_Names(file, tableorcolumn, alpha):
    file = file.replace(".csv", "")
    file = file.lower()
    file = file.replace(" ", "_")
    file = file.replace("-", "_")
    file = file.replace("%", "pct")
    file = file.replace("/", "_")
    if file.startswith(tuple(alpha)) == False:
        file = "t_" + file if tableorcolumn == "t" else "c_" + file
    return file


if __name__ == '__main__':
    createstagingdirectoty()
    download_Hospital_Zip_File()
    unzip_hospital_file()
    csvToDb()
    download_Hospital_Ranking_xlsx()
    insert_Records_Into_Temp_Rank_Focus()



